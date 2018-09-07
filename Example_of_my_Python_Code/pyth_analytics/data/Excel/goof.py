# -*- coding: utf-8 -*-
"""
Created on Tue May  1 17:21:53 2018

@author: staie

def func (*args):
    for x in args:
        print (x)
    print (sum(args))
    pass

func(1,2,3,4,5,6,7)
"""

import openpyxl


wb = openpyxl.Workbook()

sheet = wb['Sheet']

#column A contains the names of hte students
#columns B throuhg K contain the scores
#column L contains the averages
#column M contains the grade

#row 1 contains the due dates
for s in range(2,22):
    #sheet.cell(row = s, column = 12).value = '= AVERAGE(B'+str(s) +':K2'+str(s)+')'
    sheet['L' + str(s)].value = '= AVERAGE(B'+str(s) +':K2'+str(s)+')'


print (wb['Sheet'].max_row)

for s in range(20,13,-1):
    sheet.cell(row =s, column = 2).value = s + 10
    
wb.save('gradebook1.xlsx')


