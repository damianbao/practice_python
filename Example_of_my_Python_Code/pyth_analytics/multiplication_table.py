# -*- coding: utf-8 -*-
"""
Created on Tue May  1 15:44:59 2018

@author: staie
"""
import openpyxl


n = int(input('n = '))
wb = openpyxl.Workbook()

table = wb ['Sheet']

for i in range(n+1):
    table.cell(row=1, column =i+1).value = i
    
for i in range (n+1):
    table.cell(row=i+1, column=1).value = i
    
for i in range (2, n+2):
    for j in range (2, n+2):    
        table.cell(row=i, column=j).value = table.cell(row=1, column=j).value * \
               table.cell(row=i, column=1).value
        
wb.save ('Data/Excel/mult.xlsx')
