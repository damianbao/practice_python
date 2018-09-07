# -*- coding: utf-8 -*-
"""
Created on Tue Apr 24 17:22:15 2018

@author: staie
"""

import openpyxl
DATA_PATH = 'data/excel/'

wb = openpyxl.load_workbook('data/excel/')
#everytime you open a spreadsheet, save it as a new filename 
#this helps avoid corrupting originals with programming errors
wb.save('data/excel/sample2.xlsx')


summary = wb['Summary']
"""

print (header)

header = header [0]
for c in header:
    print(c.value)
    
for r in range(2, summary.max_row+1, 5):
    print (summary.cell (row=r, column = 1).value)
    """
    
summary ['C3'].value = 'hello'

summary.cell(row=3, column = 4).value = 'world' 
wb.save('data/excel/sample2.xlsx')

sheet1 = wb['Data1']
sheet2 = wb['Data2']
sheet3 = wb['Data3']

z = [1,2,3,4,5,6]
y = [2,3,4,5,6,7]
for c in range(len(z)):
    sheet1.cell(row = 1, column = c+1).value = z[c]
    sheet1.cell(row = 1, column = c+1).number_format = '0.00E+00'
    
for r in range(len(z)):
    sheet2.cell(row = r+1, column = 3).value = z[r]
    
wb.save(DATA_PATH + 'sample1.xlsx')  #'..' means go back up a level